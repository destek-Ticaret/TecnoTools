"""Excel (xlsx) export endpoint'leri.

Admin için: siparişler, ürünler, müşteriler, iadeler. openpyxl ile in-memory üretim.
"""
import io
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import require_editor
from app.models import Customer, Order, Product, ReturnRequest, User

router = APIRouter(prefix="/api/exports", tags=["exports"])


def _make_workbook() -> Workbook:
    wb = Workbook()
    return wb


def _style_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")


def _stream(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/orders.xlsx")
async def export_orders(db: AsyncSession = Depends(get_db), _: User = Depends(require_editor)):
    rows = (await db.execute(select(Order).order_by(Order.created_at.desc()))).scalars().unique().all()
    wb = _make_workbook()
    ws = wb.active
    ws.title = "Siparişler"
    _style_header(ws, [
        "Sipariş No", "Tarih", "Müşteri", "E-posta", "Telefon", "Şehir",
        "Ara Toplam", "İndirim", "Kupon", "KDV", "Kargo", "Toplam",
        "Durum", "Ödeme", "Ödeme Yöntemi", "Takip No",
    ])
    for o in rows:
        ws.append([
            o.order_no,
            (o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else ""),
            o.customer_name, o.customer_email, o.customer_phone, o.customer_city or "",
            float(o.subtotal), float(o.discount), o.coupon_code or "",
            float(o.tax), float(o.shipping), float(o.total),
            o.status, o.payment_status, o.payment_method or "", o.tracking_no or "",
        ])
    for col_letter, width in zip("ABCDEFGHIJKLMNOP", [14, 18, 22, 26, 16, 14, 12, 12, 12, 12, 12, 14, 14, 14, 14, 16]):
        ws.column_dimensions[col_letter].width = width
    return _stream(wb, f"siparisler-{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


@router.get("/products.xlsx")
async def export_products(db: AsyncSession = Depends(get_db), _: User = Depends(require_editor)):
    rows = (await db.execute(select(Product).order_by(Product.id))).scalars().unique().all()
    wb = _make_workbook()
    ws = wb.active
    ws.title = "Ürünler"
    _style_header(ws, ["ID", "Ad", "Alt başlık", "Kategori", "Fiyat", "Eski Fiyat", "Maliyet", "Kâr Marjı %", "Stok", "Aktif", "Puan", "Yorum"])
    for p in rows:
        price = float(p.price or 0)
        cost = float(p.cost) if p.cost is not None else None
        margin = round((price - cost) / price * 100, 1) if cost is not None and price > 0 else None
        ws.append([
            p.id, p.name, p.sub or "",
            (p.category.name if p.category else ""),
            price, (float(p.old_price) if p.old_price else None),
            cost, margin,
            p.stock, "Evet" if p.is_active else "Hayır",
            float(p.rating or 0), p.review_count or 0,
        ])
    return _stream(wb, f"urunler-{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


@router.get("/customers.xlsx")
async def export_customers(db: AsyncSession = Depends(get_db), _: User = Depends(require_editor)):
    rows = (await db.execute(select(Customer).order_by(Customer.id.desc()))).scalars().all()
    wb = _make_workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    _style_header(ws, ["ID", "Ad", "E-posta", "Telefon", "Şehir", "Adres", "Kayıt"])
    for c in rows:
        ws.append([
            c.id, c.name, c.email, c.phone or "", c.city or "", c.address or "",
            (c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""),
        ])
    return _stream(wb, f"musteriler-{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


@router.get("/returns.xlsx")
async def export_returns(db: AsyncSession = Depends(get_db), _: User = Depends(require_editor)):
    rows = (await db.execute(select(ReturnRequest).order_by(ReturnRequest.id.desc()))).scalars().all()
    wb = _make_workbook()
    ws = wb.active
    ws.title = "İadeler"
    _style_header(ws, ["İade #", "Sipariş", "Müşteri", "E-posta", "Sebep", "Tutar", "Durum", "Tarih"])
    for r in rows:
        ws.append([
            r.id, r.order_no, r.customer_name, r.customer_email, r.reason,
            float(r.refund_amount), r.status,
            (r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""),
        ])
    return _stream(wb, f"iadeler-{datetime.utcnow().strftime('%Y%m%d')}.xlsx")
