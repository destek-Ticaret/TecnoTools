"""Toplu ürün içe aktarma — CSV / Excel (.xlsx).

GET  /api/imports/products/template.xlsx  → boş şablon (başlıklarla)
POST /api/imports/products                → dosya yükle, ürünleri ekle/güncelle
     ?dry_run=true ile sadece önizleme (DB değişmez)

Eşleştirme: satırda `id` varsa o ürün güncellenir; yoksa `name` ile eşleşen
aktif ürün güncellenir, eşleşme yoksa yeni ürün oluşturulur. Kategori `category`
sütunundaki ada göre eşleştirilir (yoksa boş bırakılır, uyarı eklenir).
"""

import csv
import io

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import require_permission
from app.models import AuditLog, Category, Product, StockMovement, User
from app.rate_limit import limiter

router = APIRouter(prefix="/api/imports", tags=["imports"])
_can_import = require_permission("products.import")

# Şablon / kabul edilen sütun başlıkları (sıra önemli değil; başlık adıyla eşleşir)
COLUMNS = [
    "id",
    "name",
    "sub",
    "description",
    "icon",
    "category",
    "price",
    "old_price",
    "cost",
    "stock",
    "features",
    "images",
    "is_active",
]
MAX_BYTES = 5 * 1024 * 1024
MAX_ROWS = 5000


@router.get("/products/template.xlsx")
async def download_template(_: User = Depends(_can_import)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Urunler"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    # Örnek satır (kullanıcıya format ipucu)
    ws.append(
        [
            "",
            "Örnek Ürün",
            "Alt başlık",
            "Açıklama",
            "📦",
            "Aksesuar",
            199.90,
            249.90,
            120,
            50,
            "Özellik 1|Özellik 2",
            "https://...jpg|https://...jpg",
            "evet",
        ]
    )
    for col, width in zip("ABCDEFGHIJKLM", [6, 26, 20, 30, 6, 16, 10, 10, 10, 8, 28, 34, 8]):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="urun-import-sablonu.xlsx"'},
    )


def _to_bool(v) -> bool:
    s = str(v).strip().lower()
    return s in ("1", "true", "evet", "yes", "aktif", "x", "✓")


def _to_float(v):
    if v is None or str(v).strip() == "":
        return None
    return float(str(v).replace(",", ".").replace("₺", "").strip())


def _to_list(v):
    if v is None or str(v).strip() == "":
        return None
    return [s.strip() for s in str(v).split("|") if s.strip()]


def _parse_rows(raw: bytes, filename: str) -> list[dict]:
    """xlsx veya csv'yi başlık-eşlemeli dict satırlarına çevir."""
    name = (filename or "").lower()
    rows: list[list] = []
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        rows = [r for r in csv.reader(io.StringIO(text))]
    elif name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
    else:
        raise HTTPException(status_code=400, detail="Yalnız .xlsx veya .csv kabul edilir")
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return out


@router.post("/products")
@limiter.limit("10/minute")
async def import_products(
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_can_import),
):
    raw = await file.read()
    if len(raw) > MAX_BYTES:
        raise HTTPException(
            status_code=413, detail=f"Dosya {MAX_BYTES // (1024 * 1024)} MB sınırını aşıyor"
        )
    rows = _parse_rows(raw, file.filename or "")
    if not rows:
        raise HTTPException(status_code=400, detail="Dosyada veri bulunamadı")
    if len(rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"En fazla {MAX_ROWS} satır işlenebilir")

    # Kategori adı → id haritası
    cats = (await db.execute(select(Category))).scalars().all()
    cat_by_name = {c.name.strip().lower(): c.id for c in cats}

    created = updated = 0
    errors: list[dict] = []
    movements: list[StockMovement] = []

    for idx, row in enumerate(rows, start=2):  # 2: başlık + 1
        try:
            name_val = str(row.get("name") or "").strip()
            if not name_val:
                errors.append({"row": idx, "error": "name boş"})
                continue
            price = _to_float(row.get("price"))
            if price is None:
                errors.append({"row": idx, "error": "price boş/geçersiz"})
                continue

            cat_id = None
            cat_name = str(row.get("category") or "").strip()
            if cat_name:
                cat_id = cat_by_name.get(cat_name.lower())
                if cat_id is None:
                    errors.append(
                        {"row": idx, "error": f"kategori bulunamadı: {cat_name}", "warning": True}
                    )

            # is_active: dosyada sütun varsa kullan, yoksa mevcut değeri koru (yeni ürün: True)
            is_active_raw = row.get("is_active")
            is_active_present = is_active_raw is not None and str(is_active_raw).strip() != ""
            fields = dict(
                name=name_val,
                sub=(str(row.get("sub")).strip() if row.get("sub") else None),
                description=(
                    str(row.get("description")).strip() if row.get("description") else None
                ),
                icon=(str(row.get("icon")).strip() if row.get("icon") else "📦"),
                category_id=cat_id,
                price=price,
                old_price=_to_float(row.get("old_price")),
                cost=_to_float(row.get("cost")),
                stock=int(_to_float(row.get("stock")) or 0),
                features=_to_list(row.get("features")),
                images=_to_list(row.get("images")),
                is_active=_to_bool(is_active_raw) if is_active_present else True,
            )
            # Update sırasında is_active sütunu dosyada yoksa target'ın değerini koru
            is_active_in_file = is_active_present

            # Eşleştirme
            target = None
            rid = row.get("id")
            if rid is not None and str(rid).strip() != "":
                target = (
                    await db.execute(select(Product).where(Product.id == int(float(rid))))
                ).scalar_one_or_none()
                if not target:
                    errors.append({"row": idx, "error": f"id={rid} bulunamadı"})
                    continue
            else:
                target = (
                    await db.execute(select(Product).where(Product.name == name_val))
                ).scalar_one_or_none()

            if target:
                old_stock = target.stock or 0
                for k, v in fields.items():
                    # is_active dosyada yoksa mevcut değeri koru
                    if k == "is_active" and not is_active_in_file:
                        continue
                    setattr(target, k, v)
                if not dry_run and (target.stock or 0) != old_stock:
                    movements.append(
                        StockMovement(
                            product_id=target.id,
                            product_name=target.name,
                            delta=(target.stock or 0) - old_stock,
                            reason="import",
                        )
                    )
                updated += 1
            else:
                p = Product(**fields)
                if not dry_run:
                    db.add(p)
                    await db.flush()
                    if p.stock > 0:
                        movements.append(
                            StockMovement(
                                product_id=p.id, product_name=p.name, delta=p.stock, reason="import"
                            )
                        )
                created += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    hard_errors = [e for e in errors if not e.get("warning")]

    if dry_run:
        await db.rollback()
        return {
            "dry_run": True,
            "created": created,
            "updated": updated,
            "errors": errors,
            "hard_error_count": len(hard_errors),
        }

    for m in movements:
        db.add(m)
    db.add(
        AuditLog(
            actor=user.username,
            action="product-import",
            message=f"Toplu içe aktarma: {created} yeni, {updated} güncelleme, {len(hard_errors)} hata",
        )
    )
    await db.commit()
    return {"dry_run": False, "created": created, "updated": updated, "errors": errors}
