"""Toplu ürün içe aktarma — CSV + XLSX, dry-run, kategori eşleme, idempotent güncelleme."""
import io

import openpyxl


def _build_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_import_template_download(auth_client):
    """Şablon XLSX indirilir, başlıklar içerir."""
    r = await auth_client.get("/api/imports/products/template.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "name" in headers
    assert "price" in headers


async def test_import_template_requires_permission(client):
    r = await client.get("/api/imports/products/template.xlsx")
    assert r.status_code == 401


async def test_import_csv_dry_run(auth_client):
    """dry_run=true → DB değişmez, sadece özet döner."""
    csv_data = (
        b"name,price,stock\n"
        b"Yeni Urun,99.90,10\n"
    )
    files = {"file": ("products.csv", io.BytesIO(csv_data), "text/csv")}
    r = await auth_client.post("/api/imports/products?dry_run=true", files=files)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["dry_run"] is True
    assert body["created"] == 1
    assert body["updated"] == 0


async def test_import_xlsx_creates_products(auth_client, db_session):
    from app.models import Product
    from sqlalchemy import func, select

    headers = ["name", "price", "stock", "is_active"]
    data = [
        headers,
        ["İthal Matkap", "199.90", 5, "evet"],
        ["Tornavida Seti", "89.50", 12, "1"],
    ]
    files = {"file": ("urunler.xlsx", io.BytesIO(_build_xlsx(data)),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await auth_client.post("/api/imports/products", files=files)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 2

    n = (await db_session.execute(select(func.count()).select_from(Product))).scalar()
    assert n == 2


async def test_import_updates_existing_by_id(auth_client, db_session):
    from app.models import Product
    from sqlalchemy import select

    p = Product(name="Eski Ad", price=10.0, stock=1, is_active=True)
    db_session.add(p)
    await db_session.commit()
    await db_session.refresh(p)

    headers = ["id", "name", "price", "stock"]
    data = [
        headers,
        [str(p.id), "Yeni Ad", "25.0", 50],
    ]
    files = {"file": ("update.xlsx", io.BytesIO(_build_xlsx(data)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await auth_client.post("/api/imports/products", files=files)
    assert r.status_code == 200
    body = r.json()
    assert body["updated"] == 1
    assert body["created"] == 0

    pub = await auth_client.get(f"/api/products/{p.id}")
    assert pub.json()["name"] == "Yeni Ad"
    assert pub.json()["price"] == 25.0


async def test_import_updates_by_name(auth_client, db_session):
    """id yoksa name ile eşleşip günceller."""
    from app.models import Product
    p = Product(name="Aynı İsim", price=10.0, stock=1, is_active=True)
    db_session.add(p)
    await db_session.commit()

    headers = ["name", "price", "stock"]
    data = [headers, ["Aynı İsim", "30.0", 20]]
    files = {"file": ("upd.xlsx", io.BytesIO(_build_xlsx(data)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await auth_client.post("/api/imports/products", files=files)
    assert r.json()["updated"] == 1
    assert r.json()["created"] == 0


async def test_import_rejects_unsupported_extension(auth_client):
    files = {"file": ("data.json", io.BytesIO(b'{"x":1}'), "application/json")}
    r = await auth_client.post("/api/imports/products", files=files)
    assert r.status_code == 400


async def test_import_rejects_empty_file(auth_client):
    files = {"file": ("empty.csv", io.BytesIO(b""), "text/csv")}
    r = await auth_client.post("/api/imports/products", files=files)
    assert r.status_code == 400


async def test_import_rejects_oversize(auth_client):
    big = b"x" * (5 * 1024 * 1024 + 1024)  # 5 MB + 1 KB
    files = {"file": ("big.csv", io.BytesIO(big), "text/csv")}
    r = await auth_client.post("/api/imports/products", files=files)
    assert r.status_code == 413


async def test_import_handles_missing_name_and_price(auth_client):
    # name bos + price gecersiz -> 2 hata satiri
    csv_data = b"name,price,stock\n,99.90,10\nX,abc,5\n"
    files = {"file": ("bad.csv", io.BytesIO(csv_data), "text/csv")}
    r = await auth_client.post("/api/imports/products", files=files)
    assert r.status_code == 200
    body = r.json()
    assert body["created"] == 0
    assert len(body["errors"]) == 2


async def test_import_warns_on_unknown_category(auth_client):
    """Var olmayan kategori → uyarı (warning), satır yine de işlenir."""
    headers = ["name", "price", "stock", "category"]
    data = [
        headers,
        ["X", "10.0", 5, "Olmayan Kategori"],
    ]
    files = {"file": ("warn.xlsx", io.BytesIO(_build_xlsx(data)),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await auth_client.post("/api/imports/products", files=files)
    body = r.json()
    assert body["created"] == 1
    # Warning kaydı var
    assert any(e.get("warning") for e in body["errors"])
